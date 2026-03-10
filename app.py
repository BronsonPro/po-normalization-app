import os
import importlib.util
import streamlit as st
import pandas as pd
import tempfile
import re
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText

# ================== BASE DIR ==================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ================== PARTY ==================
party = st.selectbox("Select Party", ["Nykaa", "Zepto", "TiraBeauty", "TataCliq", "BlinkIt", "Scootsy", "BigBasket", "Manash", "DMart", "Myntra", "Health & Glow", "Slikk"])

# ================== PARSER LOADING ==================
convert_pdf_to_excel = None

parser_files = {
    "Nykaa": "parsers/nykaa.py",
    "Zepto": "parsers/zepto.py",
    "TiraBeauty": "parsers/tira.py",
    "TataCliq": "parsers/tatacliq.py",
    "BlinkIt": "parsers/blinkit.py",
    "Scootsy": "parsers/scootsy.py",
    "BigBasket": "parsers/bigbasket.py",
    "Manash": "parsers/manash.py",
    "DMart": "parsers/dmart.py",
    "Myntra": "parsers/myntra.py",
    "Health & Glow": "parsers/healthandglow.py",
    "Slikk": "parsers/slikk.py"
}

parser_path = os.path.join(BASE_DIR, parser_files.get(party, ""))

if os.path.exists(parser_path):
    spec = importlib.util.spec_from_file_location("party_parser", parser_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    convert_pdf_to_excel = getattr(module, "convert_pdf_to_excel", None)


    # ================== PARTY CODE MASTER ==================
    PARTY_CODE_FILE = os.path.join(BASE_DIR, "PartyCode.xlsx")

    def load_party_code_master():

        if not os.path.exists(PARTY_CODE_FILE):
            return None

        pc = pd.read_excel(PARTY_CODE_FILE)
        pc.columns = pc.columns.astype(str).str.strip()

        rename = {}

        for c in pc.columns:
            cl = str(c).lower()

            if "party" in cl and "name" in cl:
                rename[c] = "Party Name"

            elif "pin" in cl or "zip" in cl:
                rename[c] = "Pincode"

            elif "party" in cl and "code" in cl:
                rename[c] = "Party Code"

        pc = pc.rename(columns=rename)

        required = ["Party Name", "Pincode", "Party Code"]
        if not all(c in pc.columns for c in required):
            return None

        pc["Party Name"] = pc["Party Name"].astype(str).str.strip().str.lower()
        pc["Pincode"] = pc["Pincode"].astype(str).str.extract(r"(\d{6})")

        return pc[required]


# ================== EMAIL CONFIGURATION ==================
def load_email_config():
    """Load email configuration from Streamlit secrets or Excel file"""
    try:
        # Try Streamlit secrets first (secure)
        if 'email' in st.secrets:
            config = {
                'Sender_Email': st.secrets['email']['Sender_Email'],
                'Sender_Password': st.secrets['email']['Sender_Password'],
                'Recipient_Email': st.secrets['email']['Recipient_Email'],
                'SMTP_Server': st.secrets['email']['SMTP_Server'],
                'SMTP_Port': st.secrets['email']['SMTP_Port']
            }
            
            # Add Django token if available
            if 'django' in st.secrets:
                config['Django_Token'] = st.secrets['django']['Django_Token']
            
            return config
    except Exception as e:
        print(f"Error loading secrets: {e}")
    
    # Fallback to Email_Config.xlsx (if exists)
    try:
        config_path = os.path.join(BASE_DIR, "Email_Config.xlsx")
        if os.path.exists(config_path):
            config_df = pd.read_excel(config_path)
            config = {}
            for _, row in config_df.iterrows():
                config[row['Setting']] = row['Value']
            return config
    except Exception as e:
        st.error(f"Error loading email config: {e}")
    
    return None


def send_email_with_attachment(file_path, po_number, party_name):
    """Send email with PO file attached"""
    
    # Load configuration
    config = load_email_config()
    if not config:
        return False, "Email configuration not found"
    
    sender_email = config.get('Sender_Email')
    sender_password = config.get('Sender_Password')
    recipient_email = config.get('Recipient_Email')
    smtp_server = config.get('SMTP_Server', 'smtp.gmail.com')
    smtp_port = config.get('SMTP_Port', '587')
    
    # Convert port to integer
    try:
        smtp_port = int(smtp_port)
    except:
        smtp_port = 587
    
    if not all([sender_email, sender_password, recipient_email]):
        missing = []
        if not sender_email:
            missing.append("Sender_Email")
        if not sender_password:
            missing.append("Sender_Password")
        if not recipient_email:
            missing.append("Recipient_Email")
        return False, f"Email configuration incomplete. Missing: {', '.join(missing)}"
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = f"PO {po_number} - {party_name}"
        
        # Email body
        body = f"""Hello,

Please find attached the processed Purchase Order.

PO Number: {po_number}
Party: {party_name}
Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}

This is an automated email from PO Processing System.

Best regards,
PO Automation"""
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach file
        filename = os.path.basename(file_path)
        
        with open(file_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename= {filename}')
        msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, recipient_email, text)
        server.quit()
        
        return True, f"Email sent successfully to {recipient_email}!"
    
    except smtplib.SMTPAuthenticationError as e:
        return False, f"Authentication failed. Check your email/password. Error: {str(e)}"
    
    except smtplib.SMTPException as e:
        return False, f"SMTP error occurred: {str(e)}"
    
    except Exception as e:
        return False, f"Email failed: {str(e)}"


# ================== DJANGO UPLOAD ==================
def upload_to_django(po_number, party_code_value, po_date, po_expiry_date):
    """Upload PO data to Django API"""
    
    try:
        # Get already-processed dataframe from session state
        if 'upd_df' not in st.session_state:
            return False, "Product data not found. Please run validation again."
        
        df = st.session_state['upd_df'].copy()

        # Fix EAN - convert to string and clean
        if 'EAN' in df.columns:
            df['EAN'] = df['EAN'].astype(str).str.strip().str.replace('.0', '', regex=False)
            df = df[~df['EAN'].isin(['', 'nan', 'NaN', 'None'])]

        # Find EAN and Quantity columns
        ean_col = None
        qty_col = None
        for c in df.columns:
            cl = c.lower()
            if "ean" in cl:
                ean_col = c
            if "qty" in cl or "quantity" in cl:
                qty_col = c

        if not ean_col:
            return False, f"EAN column not found. Columns: {df.columns.tolist()}"
        if not qty_col:
            return False, f"Quantity column not found. Columns: {df.columns.tolist()}"

        # Filter valid rows
        df = df[df[qty_col].apply(lambda x: float(str(x)) > 0 if str(x) not in ['', 'nan'] else False)]

        if df.empty:
            return False, "No valid product rows found"

        # Date formatter
        def format_date(date_str, fallback=""):
            from datetime import datetime
            date_str = str(date_str).strip()
            if not date_str or date_str == "nan":
                return fallback
            date_str = re.sub(r',?\s+\d{1,2}:\d{2}.*$', '', date_str).strip()
            date_str = re.sub(r'([A-Za-z])\.', r'\1', date_str).strip()
            for fmt in ["%d/%b/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y",
                        "%m/%d/%Y", "%d.%m.%Y", "%b %d, %Y", "%B %d, %Y",
                        "%d %b %Y", "%d %B %Y"]:
                try:
                    return datetime.strptime(date_str, fmt).strftime("%d-%m-%Y")
                except:
                    continue
            return fallback

        formatted_po_date = format_date(po_date)
        # If expiry date missing, default to 60 days from PO date
        from datetime import datetime, timedelta
        if not po_expiry_date or str(po_expiry_date).strip() in ["", "nan"]:
            try:
                base = datetime.strptime(formatted_po_date, "%d-%m-%Y")
                formatted_expiry_date = (base + timedelta(days=60)).strftime("%d-%m-%Y")
            except:
                formatted_expiry_date = formatted_po_date
        else:
            formatted_expiry_date = format_date(po_expiry_date)

        # Build JSON payload
        payload = []
        skipped = 0

        for _, row in df.iterrows():
            try:
                ean_val = str(row[ean_col]).strip().replace(".0", "").replace(" ", "")

                if not ean_val or ean_val == "nan":
                    skipped += 1
                    continue

                try:
                    qty_int = int(float(str(row[qty_col])))
                except:
                    skipped += 1
                    continue

                if qty_int <= 0:
                    skipped += 1
                    continue

                # MRP - use Base Rate if MRP is 0 (e.g. TataCliq)
                try:
                    mrp_val = round(float(str(row.get("MRP", 0))), 2)
                    if mrp_val == 0 and "Base Rate" in df.columns:
                        mrp_val = round(float(str(row.get("Base Rate", 0))), 2)
                except:
                    mrp_val = 0

                try:
                    base_rate_val = round(float(str(row.get("Base Rate", 0))), 2)
                except:
                    base_rate_val = 0

                try:
                    gst_val = round(float(str(row.get("GST %", 0))), 2)
                except:
                    gst_val = 0

                try:
                    total_val = round(float(str(row.get("Total", 0))), 2)
                except:
                    total_val = 0

                try:
                    hsn_val = str(row.get("HSN Code", "")).strip().replace(".0", "")
                except:
                    hsn_val = ""

                item = {
                    "po_no": str(po_number).strip(),
                    "customer_code": str(party_code_value).strip(),
                    "barcode": ean_val,
                    "quantity": qty_int,
                    "po_date": formatted_po_date,
                    "po_expiry_date": formatted_expiry_date,
                    "mrp": mrp_val,
                    "base_rate": base_rate_val,
                    "gst_percentage": gst_val,
                    "total_amount": total_val,
                    "hsn_code": hsn_val
                }
                payload.append(item)

            except Exception as e:
                skipped += 1
                continue

        if not payload:
            return False, f"No valid items to upload. {skipped} rows skipped."

        # Load JWT token from config or secrets
        config = load_email_config()
        django_token = ""
        
        # Try loading from config first
        if config:
            raw_token = config.get('Django_Token', '')
            if raw_token and str(raw_token).strip() not in ['', 'nan']:
                django_token = str(raw_token).strip()
        
        # If not found, try secrets directly
        if not django_token:
            try:
                if 'django' in st.secrets:
                    django_token = st.secrets['django']['Django_Token']
            except:
                pass
        
        if not django_token:
            return False, "Django token not found. Add Django_Token to secrets or Email_Config.xlsx"

        response = requests.post(
            "http://16.16.170.252/packingorder/purchase-orders/import/",
            json=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {django_token}"
            },
            timeout=30
        )

        if response.status_code == 201:
            return True, f"✅ Successfully uploaded {len(payload)} items to mobile app!"
        elif response.status_code == 207:
            return True, f"⚠️ Partial success: {response.text[:200]}"
        else:
            return False, f"Upload failed. Status: {response.status_code}. Response: {response.text[:200]}"

    except requests.exceptions.ConnectionError:
        return False, "Connection failed. Check if Django server is running."
    except requests.exceptions.Timeout:
        return False, "Request timed out."
    except Exception as e:
        return False, f"Upload failed: {str(e)}"


# ================== RACK MASTER ==================
RACK_FILE_PATH = os.path.join(BASE_DIR, "Rack number.xlsx")

def load_rack_master():
    if not os.path.exists(RACK_FILE_PATH):
        return None

    rack = pd.read_excel(RACK_FILE_PATH)
    rack.columns = rack.columns.astype(str).str.strip()

    rename = {}
    for c in rack.columns:
        cl = str(c).lower()
        if "ean" in cl:
            rename[c] = "EAN"
        if "rack" in cl:
            rename[c] = "Rack Number"

    rack = rack.rename(columns=rename)

    if "EAN" not in rack.columns or "Rack Number" not in rack.columns:
        return None

    rack["EAN"] = pd.to_numeric(rack["EAN"], errors="coerce")
    rack = rack.dropna(subset=["EAN"])
    rack["EAN"] = rack["EAN"].astype("int64")

    return rack[["EAN", "Rack Number"]]


# ================== TITLE ==================
st.title("📄 PO Normalization & Validation App")


# ================== UTIL ==================
def read_normalized_po_table(excel_path):
    raw = pd.read_excel(excel_path, header=None)

    header_row = None
    for i in range(len(raw)):
        row = raw.iloc[i].astype(str).str.lower().tolist()
        if any("ean" in str(cell) for cell in row):
            header_row = i
            break

    if header_row is None:
        raise Exception("Could not locate product table (EAN not found)")

    df = pd.read_excel(excel_path, header=header_row)
    df.columns = df.columns.astype(str).str.strip()

    # Stop at total row
    stop_words = ["total amount", "grand total", "total tax"]
    for i in range(len(df)):
        row_text = " ".join(df.iloc[i].astype(str).str.lower().tolist())
        if any(word in row_text for word in stop_words):
            df = df.iloc[:i]
            break

    # Rename EAN column variations
    for c in df.columns:
        if "ean" in c.lower() and c != "EAN":
            df = df.rename(columns={c: "EAN"})
            break

    # Remove DC / LOCATION rows
    df["__ean_num"] = pd.to_numeric(df["EAN"], errors="coerce")

    qty_col = None
    for c in df.columns:
        if "qty" in str(c).lower() or "quantity" in str(c).lower():
            qty_col = c
            break

    if qty_col:
        df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)

    numeric_cols = []
    for c in df.columns:
        if any(k in c.lower() for k in ["mrp", "base", "rate", "amount", "total"]):
            numeric_cols.append(c)

    for c in numeric_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # SCOOTSY FIX #1: Skip EAN validation for Scootsy (they use Item Code)
    if party == "Scootsy":
        if qty_col:
            df = df[df[qty_col] > 0].copy()
    else:
        df = df[
            (df["__ean_num"].notna()) &
            ((df[qty_col] > 0) if qty_col else True) &
            ~((df[numeric_cols].sum(axis=1)) == 0)
        ].copy()
    
    df.drop(columns="__ean_num", inplace=True)
    if "EAN" in df.columns:
        df = df.drop_duplicates(subset=["EAN"], keep="first")
    df.reset_index(drop=True, inplace=True)

    return df, raw, header_row


def format_2_dec(x):
    try:
        return f"{float(x):.2f}"
    except:
        return x


# ================== STEP 1 ==================
st.header("Step 1: Upload PO")

po_file = st.file_uploader("Upload PO (PDF / Excel)", type=["pdf", "xlsx", "xls"])

po_df = None
raw_po = None
table_header_row = None
converted_po_path = None

if po_file:
    st.success(f"Loaded: {po_file.name}")
    ext = po_file.name.split(".")[-1].lower()

    if ext == "pdf":
        if convert_pdf_to_excel is None:
            st.error("Parser not available for selected party.")
            st.stop()

        input_path = os.path.join(tempfile.gettempdir(), po_file.name)
        converted_po_path = os.path.join(tempfile.gettempdir(), "po_converted.xlsx")

        with open(input_path, "wb") as f:
            f.write(po_file.read())

        convert_pdf_to_excel(input_path, converted_po_path)
        po_df, raw_po, table_header_row = read_normalized_po_table(converted_po_path)
        st.download_button("⬇ Download Converted PO", open(converted_po_path, "rb"), "PO_Converted.xlsx")

    else:
        if convert_pdf_to_excel is not None:
            input_path = os.path.join(tempfile.gettempdir(), po_file.name)
            converted_po_path = os.path.join(tempfile.gettempdir(), "po_converted.xlsx")
            with open(input_path, "wb") as f:
                f.write(po_file.read())
            convert_pdf_to_excel(input_path, converted_po_path)
        else:
            converted_po_path = po_file

        po_df, raw_po, table_header_row = read_normalized_po_table(converted_po_path)
        st.download_button("⬇ Download Converted PO", open(converted_po_path, "rb"), "PO_Converted.xlsx")


# ================== STEP 2 ==================
st.header("Step 2: Upload Master File")

master_file = st.file_uploader("Upload Master Excel", type=["xlsx", "xls"], key="master")
master_df = None

if master_file:
    master_df = pd.read_excel(master_file)
    master_df.columns = master_df.columns.astype(str).str.strip()


# ================== STEP 3 ==================
st.header("Step 3: Validate")

# Show persistent success message if validation passed
if st.session_state.get('validation_success', False):
    st.success("✅ No mismatches found. Updating Product Name & HSN Code from Master.")

if po_df is not None and master_df is not None:
    if st.button("▶ Run Validation"):
        po = po_df.copy()
        master = master_df.copy()
        rack_master = load_rack_master()
        party_code_master = load_party_code_master()

        # SMART RENAME FUNCTION
        def normalize(df, is_master=False):
            rename = {}
            for c in df.columns:
                cl = str(c).strip().lower()

                # SCOOTSY FIX #2: Handle Item Code FIRST (before EAN)
                if "item code" in cl or "item_code" in cl:
                    rename[c] = "Item Code"
                elif "ean" in cl or "upc" in cl or "brand sku code" in cl:
                    rename[c] = "EAN"
                elif "mrp" in cl:
                    rename[c] = "MRP"
                elif "gst" in cl or ("tax" in cl and "rate" in cl):
                    if "landing" not in cl and "incl" not in cl and "excl" not in cl and "taxable" not in cl:
                        if not (party == "BigBasket" and not is_master):
                            rename[c] = "GST %"
                elif "hsn" in cl:
                    rename[c] = "HSN Code"
                elif (
                    "name as per brand ean" in cl
                    or ("product" in cl and "name" in cl)
                    or ("sku" in cl and "name" in cl and "revised" not in cl)
                    or "item description" in cl
                    or "material" in cl
                    or "article description" in cl
                ):
                    rename[c] = "Product Name"

                # Party-specific renames
                if party == "Zepto":
                    if not is_master and ("base" in cl or "unit" in cl):
                        rename[c] = "Base Rate"
                    if is_master and any(k in cl for k in ["cp", "excl", "base"]):
                        rename[c] = "Base Rate"

                if party == "TiraBeauty":
                    if not is_master and any(k in cl for k in ["base", "cost"]):
                        rename[c] = "Base Rate"
                    if is_master and "taxable" in cl:
                        rename[c] = "Base Rate"

                if party == "BlinkIt":
                    if is_master and ("incl" in cl and "gst" in cl):
                        rename[c] = "Base Rate"

                if party == "TataCliq":
                    if is_master and "taxable" in cl and "amt" in cl:
                        rename[c] = "Base Rate"

                if party == "Scootsy":
                    if is_master and "taxable" in cl and "price" in cl:
                        rename[c] = "Base Rate"

                if party == "BigBasket":
                    if not is_master and "landing" in cl and "cost" in cl:
                        rename[c] = "Base Rate"
                    if is_master and "new supply rate" in cl:
                        rename[c] = "Base Rate"
                    if not is_master and cl == "gst%":
                        rename[c] = "GST %"
                    if is_master and "barcode" in cl:
                        rename[c] = "EAN"

                if party == "Manash":
                    if is_master and "taxable rate" in cl:
                        rename[c] = "Base Rate"
                    if is_master and "ean_code" in cl or "ean code" in cl:
                        rename[c] = "EAN"
                    if is_master and cl == "hsn":
                        rename[c] = "HSN Code"

                if party == "DMart":
                    if is_master and "taxable" in cl:
                        rename[c] = "Base Rate"
                    if is_master and "ean" in cl:
                        rename[c] = "EAN"
                    if is_master and "hsn" in cl:
                        rename[c] = "HSN Code"
                    if is_master and "tax rate" in cl:
                        rename[c] = "GST %"

                if party == "Myntra":
                    if is_master and "taxable cost" in cl:
                        rename[c] = "Base Rate"
                    if is_master and "ean code" in cl:
                        rename[c] = "EAN"
                    if is_master and "gst%" in cl or "gst %" in cl:
                        rename[c] = "GST %"

                if party == "Health & Glow":
                    if is_master and ("en/barcode" in cl or "barcode" in cl):
                        rename[c] = "EAN"
                    if is_master and "taxable rate" in cl:
                        rename[c] = "Base Rate"
                    if is_master and cl == "gst":
                        rename[c] = "GST %"
                    if is_master and "hsn code" in cl:
                        rename[c] = "HSN Code"

                if party == "Slikk":
                    if is_master and ("ean code" in cl or "ean" in cl):
                        rename[c] = "EAN"
                    if is_master and "taxable rate" in cl:
                        rename[c] = "Base Rate"
                    if is_master and "tax rate" in cl and "taxable" not in cl:
                        rename[c] = "GST %"
                    if is_master and "hsn code" in cl:
                        rename[c] = "HSN Code"

            df = df.rename(columns=rename)
            df = df.loc[:, ~df.columns.duplicated()]
            return df

        po = normalize(po, is_master=False)
        master = normalize(master, is_master=True)

        # SCOOTSY FIX #3: Deduplicate on Item Code for Scootsy, EAN for others
        if party == "Scootsy":
            if "Item Code" in po.columns:
                po = po.drop_duplicates(subset=["Item Code"], keep="first")
        else:
            if "EAN" in po.columns:
                po = po.drop_duplicates(subset=["EAN"], keep="first")
        
        master = master.sort_values("EAN").drop_duplicates(subset=["EAN"], keep="first")

        # Define required columns per party
        if party == "Nykaa":
            po_req = ["EAN", "MRP", "GST %"]
            master_req = ["EAN", "MRP", "GST %", "Product Name", "HSN Code"]
        elif party == "TataCliq":
            po_req = ["EAN", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "BigBasket":
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "Manash":
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "DMart":
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "Myntra":
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "Health & Glow":
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "Slikk":
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]
        elif party == "Scootsy":
            po_req = ["Item Code", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "Item Code", "Product Name"]
        else:
            po_req = ["EAN", "MRP", "Base Rate", "GST %"]
            master_req = ["EAN", "MRP", "Base Rate", "GST %", "Product Name", "HSN Code"]

        # Check required columns
        for c in po_req:
            if c not in po.columns:
                st.error(f"PO missing column: {c}")
                st.stop()

        for c in master_req:
            if c not in master.columns:
                st.error(f"Master missing column: {c}")
                st.stop()

        # Validate numeric columns
        if party == "Scootsy":
            po["Item Code"] = pd.to_numeric(po["Item Code"], errors="coerce")
            po = po.dropna(subset=["Item Code"])
            po["Item Code"] = po["Item Code"].astype("int64")

            master["Item Code"] = pd.to_numeric(master["Item Code"], errors="coerce")
            master = master.dropna(subset=["Item Code"])
            master["Item Code"] = master["Item Code"].astype("int64")

            master["EAN"] = pd.to_numeric(master["EAN"], errors="coerce")
            master = master.dropna(subset=["EAN"])
            master["EAN"] = master["EAN"].astype("int64")
        else:
            po["EAN"] = pd.to_numeric(po["EAN"], errors="coerce")
            master["EAN"] = pd.to_numeric(master["EAN"], errors="coerce")
            po = po.dropna(subset=["EAN"])
            master = master.dropna(subset=["EAN"])
            po["EAN"] = po["EAN"].astype("int64")
            master["EAN"] = master["EAN"].astype("int64")

        for c in po_req:
            if c not in ["EAN", "Item Code"]:
                po[c] = pd.to_numeric(po[c], errors="coerce").fillna(0)

        for c in master_req:
            if c not in ["EAN", "Item Code", "Product Name", "HSN Code"]:
                master[c] = pd.to_numeric(master[c], errors="coerce").fillna(0)

        master["GST %"] = master["GST %"].apply(lambda x: x * 100 if x <= 1 else x)
        po["GST %"] = po["GST %"].round(2)
        master["GST %"] = master["GST %"].round(2)

        # Validation merge
        if party == "Scootsy":
            merged = po.merge(master, on="Item Code", how="left", suffixes=("_PO", "_MASTER"))
        else:
            merged = po.merge(master, on="EAN", how="left", suffixes=("_PO", "_MASTER"))

        for col in merged.select_dtypes(include=['object']).columns:
            merged[col] = merged[col].astype(str)

        # Check for mismatches
        reasons = []
        for _, r in merged.iterrows():
            issue = []
            if party != "TataCliq":
                if abs(r["MRP_PO"] - r["MRP_MASTER"]) > 0.01:
                    issue.append("MRP Mismatch")
            if party != "Nykaa":
                if abs(r["Base Rate_PO"] - r["Base Rate_MASTER"]) > 0.01:
                    issue.append("Base Rate Mismatch")
            if abs(r["GST %_PO"] - r["GST %_MASTER"]) > 0.01:
                issue.append("GST % Mismatch")
            reasons.append(", ".join(issue))

        merged["Reason"] = reasons
        mismatch = merged[merged["Reason"] != ""]

        if not mismatch.empty:
            if party == "Nykaa":
                report = mismatch[["EAN", "MRP_PO", "MRP_MASTER", "GST %_PO", "GST %_MASTER", "Reason"]]
            elif party == "TataCliq":
                report = mismatch[["EAN", "Base Rate_PO", "Base Rate_MASTER", "GST %_PO", "GST %_MASTER", "Reason"]]
            else:
                report = mismatch[["EAN", "MRP_PO", "MRP_MASTER", "Base Rate_PO", "Base Rate_MASTER", "GST %_PO", "GST %_MASTER", "Reason"]]

            report = report.round(2)
            path = os.path.join(tempfile.gettempdir(), "Mismatch_Report.xlsx")
            report.to_excel(path, index=False)

            st.error("❌ Mismatch found")
            st.dataframe(report, width="stretch")
            st.download_button("⬇ Download Mismatch Report", open(path, "rb"), "Mismatch_Report.xlsx")

        else:
            st.session_state['validation_success'] = True
            st.success("✅ No mismatches found. Updating Product Name & HSN Code from Master.")

            # SCOOTSY FIX #4: Proper merge for Scootsy
            if party == "Scootsy":
                # Drop empty EAN from PO, then merge on Item Code
                if "EAN" in po.columns:
                    po = po.drop(columns=["EAN"])
                
                # DEBUG - Check Item Codes before merge
                st.write("🔍 Scootsy Merge Debug:")
                st.write(f"PO Item Codes: {sorted(po['Item Code'].tolist())}")
                st.write(f"Master has {len(master)} items")
                st.write(f"Master Item Codes (sample): {sorted(master['Item Code'].head(20).tolist())}")
                
                # Check if PO Item Codes exist in Master
                po_codes = set(po['Item Code'].tolist())
                master_codes = set(master['Item Code'].tolist())
                matching = po_codes & master_codes
                missing = po_codes - master_codes
                
                st.write(f"✅ Matching Item Codes: {sorted(matching)}")
                if missing:
                    st.write(f"❌ PO Item Codes NOT in Master: {sorted(missing)}")
                
                upd = po.merge(master[["Item Code", "EAN", "Product Name"]], on="Item Code", how="left")
                
                st.write(f"After merge: {len(upd)} rows")
                st.write(f"EAN values: {upd['EAN'].tolist()}")
                st.dataframe(upd[["Item Code", "EAN", "Product Name"]])
            else:
                # All other parties: Standard EAN merge
                upd = po.merge(
                    master[["EAN", "Product Name", "HSN Code"]],
                    on="EAN",
                    how="left",
                    suffixes=("_PO", "_MASTER")
                )

            # Add rack number
            if rack_master is not None:
                upd = upd.merge(rack_master, on="EAN", how="left")
            else:
                upd["Rack Number"] = ""

            if "Rack Number" in upd.columns:
                cols = [c for c in upd.columns if c != "Rack Number"] + ["Rack Number"]
                upd = upd[cols]

            # TataCliq: Populate MRP from master
            if party == "TataCliq":
                if "MRP" in upd.columns and "EAN" in upd.columns:
                    mrp_lookup = master.set_index("EAN")["MRP"].to_dict()
                    upd["MRP"] = upd["EAN"].map(mrp_lookup).fillna(0)

            if "Product Name_MASTER" in upd.columns:
                upd["Product Name"] = upd["Product Name_MASTER"].fillna(upd["Product Name_PO"])

            if "HSN Code_MASTER" in upd.columns:
                upd["HSN Code"] = upd["HSN Code_MASTER"].fillna(upd["HSN Code_PO"])

            upd.drop(columns=[c for c in upd.columns if c.endswith("_PO") or c.endswith("_MASTER")], inplace=True)

            # SCOOTSY FIX #5: Convert EAN properly for Scootsy (avoid scientific notation)
            if "EAN" in upd.columns:
                if party == "Scootsy":
                    upd["EAN"] = upd["EAN"].apply(lambda x: f"{int(x)}" if pd.notna(x) and x != "" and x != 0 else "")
                else:
                    upd["EAN"] = upd["EAN"].astype(str).str.replace(".0","", regex=False)

            final_raw = raw_po.copy()
            start_row = table_header_row + 1

            money_cols = [c for c in upd.columns if any(k in c.lower() for k in ["total", "value", "rate", "mrp", "amount", "base", "tax"])]

            header_values = final_raw.iloc[table_header_row].astype(str).str.strip().tolist()

            # Ensure Rack column exists
            if "Rack Number" not in header_values:
                final_raw["Rack Number"] = ""
                final_raw.at[table_header_row, "Rack Number"] = "Rack Number"

            header_values = final_raw.iloc[table_header_row].astype(str).str.strip().tolist()
            col_index_map = {h: idx for idx, h in enumerate(header_values)}

            for i in range(len(upd)):
                for col in upd.columns:
                    if col not in col_index_map:
                        continue
                    j = col_index_map[col]
                    val = upd.at[i, col]
                    final_raw.iat[start_row + i, j] = format_2_dec(val) if col in money_cols else val

            # SCOOTSY FIX #6: Remove Item Code column from final sheet
            if party == "Scootsy":
                header_values = final_raw.iloc[table_header_row].astype(str).str.strip().tolist()
                if "Item Code" in header_values:
                    item_code_idx = header_values.index("Item Code")
                    final_raw.drop(final_raw.columns[item_code_idx], axis=1, inplace=True)

            # Get PO number
            po_number = "PO"
            for i in range(table_header_row):
                row = final_raw.iloc[i].astype(str).str.lower().tolist()
                if "po no" in row or "po number" in row:
                    try:
                        po_number = str(final_raw.iloc[i, 1]).strip()
                    except:
                        po_number = "PO"
                    break

            safe_party = party.replace(" ", "").replace("/", "")
            safe_po = po_number.replace("/", "_").replace("\\", "_").replace(" ", "")
            filename = f"{safe_party}_{safe_po}.xlsx"
            final_path = os.path.join(tempfile.gettempdir(), filename)

            final_raw.to_excel(final_path, index=False, header=False)

            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
            from math import ceil

            wb = load_workbook(final_path)
            ws = wb.active

            # Add party code
            party_code_value = ""
            if party_code_master is not None:
                try:
                    party_name_sheet = ""
                    shipping_pin = ""

                    for row in ws.iter_rows(min_row=1, max_row=table_header_row):
                        label = str(row[0].value).strip().lower() if row[0].value else ""
                        
                        if label == "party name":
                            party_name_sheet = str(row[1].value).strip().lower() if row[1].value else ""
                        
                        if label == "shipping address":
                            addr = str(row[1].value) if row[1].value else ""
                            pin_match = re.findall(r"\d{6}", addr)
                            if pin_match:
                                shipping_pin = pin_match[0]
                            elif addr.strip().isdigit() and len(addr.strip()) == 6:
                                shipping_pin = addr.strip()
                            else:
                                shipping_pin = ""

                    if party_name_sheet and shipping_pin:
                        party_name_clean = re.sub(r'[^a-z0-9 ]', '', party_name_sheet.lower())
                        party_code_master["_clean_name"] = party_code_master["Party Name"].apply(
                            lambda x: re.sub(r'[^a-z0-9 ]', '', str(x).lower())
                        )

                        match = party_code_master[
                            (party_code_master["Pincode"].astype(str) == str(shipping_pin)) &
                            (party_code_master["_clean_name"] == party_name_clean)
                        ]
                        
                        if match.empty:
                            match = party_code_master[
                                (party_code_master["Pincode"].astype(str) == str(shipping_pin)) &
                                (party_code_master["_clean_name"].str.contains(party_name_clean, na=False, regex=False))
                            ]

                        if not match.empty:
                            party_code_value = str(match.iloc[0]["Party Code"])

                except Exception as e:
                    party_code_value = ""

            insert_row = None
            for r in range(1, table_header_row + 2):
                val = ws.cell(row=r, column=1).value
                if val and str(val).strip().lower() == "party name":
                    insert_row = r + 1
                    break

            if insert_row is None:
                insert_row = 2

            ws.insert_rows(insert_row)
            ws.cell(row=insert_row, column=1).value = "Party Code"
            ws.cell(row=insert_row, column=2).value = party_code_value

            # Shipping address → keep only pincode
            for row in ws.iter_rows(min_row=1, max_row=table_header_row):
                field_cell = row[0]
                if field_cell.value and str(field_cell.value).strip().lower() == "shipping address":
                    val_cell = row[1]
                    if val_cell.value:
                        addr_text = str(val_cell.value)
                        pin_match = re.search(r"\b[1-9][0-9]{5}\b", addr_text)
                        if pin_match:
                            val_cell.value = pin_match.group(0)
                        else:
                            val_cell.value = ""

            # Lock summary as text
            for row in ws.iter_rows():
                if row[0].value and str(row[0].value).strip().lower() in ["total base value", "total tax", "grand total"]:
                    row[1].value = str(row[1].value)
                    row[1].number_format = "@"

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            # Fix EAN column
            ean_col_idx = None
            for col in ws.iter_cols(min_row=table_header_row + 1, max_row=table_header_row + 1):
                if str(col[0].value).strip().lower() == "ean":
                    ean_col_idx = col[0].column
                    break

            if ean_col_idx:
                start = table_header_row + 2
                end = start + len(upd) - 1

                for r in range(start, end + 1):
                    cell = ws.cell(row=r, column=ean_col_idx)
                    if cell.value not in ("", None):
                        try:
                            cell.value = str(cell.value).split(".")[0]
                        except:
                            pass
                        cell.number_format = "@"

            # Force column widths
            fixed_widths = {
                "A": 13, "B": 18, "C": 36, "D": 12, "E": 9,
                "F": 9, "G": 11, "H": 7, "I": 12, "J": 12
            }

            for col_letter, width in fixed_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Auto row height
            for row in ws.iter_rows():
                row_idx = row[0].row
                max_lines = 1

                for cell in row:
                    if cell.value:
                        text = str(cell.value)
                        col_width = ws.column_dimensions[cell.column_letter].width or 15
                        est_lines = ceil(len(text) / (col_width * 1.1))
                        max_lines = max(max_lines, est_lines, text.count("\n") + 1)

                ws.row_dimensions[row_idx].height = max(18, max_lines * 15)
            
            # Adjust specific row heights
            for row in ws.iter_rows(min_row=1, max_row=8):
                field_cell = row[0]
                if field_cell.value and str(field_cell.value).strip().lower() == "shipping address":
                    r = field_cell.row
                    ws.row_dimensions[r].height = 45
            
            ws.row_dimensions[10].height = 30

            # Print settings
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.scale = 100
            ws.page_margins.left = 0.3
            ws.page_margins.right = 0.3
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5

            # Add gridlines
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            num_cols = ws.max_column
            
            # Header fields (Rows 1-8, Columns A-B)
            for row in range(1, 9):
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col == 1:
                        cell.font = Font(bold=True)
            
            # Table header row (Row 10, all columns)
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=10, column=col)
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Table content (Row 11 onwards)
            last_data_row = 10
            for row_idx in range(11, ws.max_row + 1):
                has_data = False
                for col in range(1, min(5, num_cols + 1)):
                    if ws.cell(row=row_idx, column=col).value:
                        has_data = True
                        break
                if has_data:
                    last_data_row = row_idx
                else:
                    break
            
            for row in range(11, last_data_row + 1):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Summary fields (3 rows, Columns A-B)
            summary_start = last_data_row + 2
            for row in range(summary_start, summary_start + 3):
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col == 1:
                        cell.font = Font(bold=True)

            wb.save(final_path)

            st.success("✅ Final PO formatted and ready!")

            # Store in session state
            st.session_state['final_path'] = final_path
            st.session_state['final_name'] = os.path.basename(final_path)
            st.session_state['po_number'] = po_number
            st.session_state['party'] = party
            st.session_state['party_code_value'] = party_code_value

            # Merge MRP from master for platforms where PO has no MRP
            if "MRP_MASTER" in merged.columns:
                upd["MRP"] = pd.to_numeric(merged["MRP_MASTER"], errors="coerce").fillna(
                    pd.to_numeric(merged.get("MRP_PO", 0), errors="coerce").fillna(0)
                )
            st.session_state['upd_df'] = upd.copy()
            
            # Extract PO Date and Expiry Date
            po_date = ""
            po_expiry_date = ""
            for i in range(table_header_row):
                row_vals = final_raw.iloc[i].astype(str).str.lower().tolist()
                if "po date" in row_vals or "po_date" in row_vals:
                    try:
                        po_date = str(final_raw.iloc[i, 1]).strip()
                    except:
                        pass
                if "po expiry date" in row_vals or "po_expiry_date" in row_vals:
                    try:
                        po_expiry_date = str(final_raw.iloc[i, 1]).strip()
                    except:
                        pass
            
            st.session_state['po_date'] = po_date
            st.session_state['po_expiry_date'] = po_expiry_date
            st.rerun()

else:
    st.info("Upload both PO and Master file to enable validation.")


# ================== DOWNLOAD & EMAIL SECTION ==================
if 'final_path' in st.session_state:
    
    st.markdown("---")
    st.subheader("📥 Download, Email & Upload")
    
    col1, col2 = st.columns(2)
    
    with col1:
        with open(st.session_state['final_path'], "rb") as file:
            st.download_button(
                "⬇ Download Final PO",
                file,
                st.session_state['final_name'],
                "application/vnd.ms-excel",
                key="download_final_po"
            )
    
    with col2:
        email_config = load_email_config()
        if email_config and email_config.get('Recipient_Email'):
            if st.button("📧 Email PO & Upload to App", key="email_and_upload_button"):
                # Email PO
                with st.spinner("Sending email..."):
                    email_success, email_message = send_email_with_attachment(
                        st.session_state['final_path'],
                        st.session_state['po_number'],
                        st.session_state['party']
                    )
                    if email_success:
                        st.success(email_message)
                    else:
                        st.error(email_message)
                
                # Upload to App
                with st.spinner("Uploading to mobile app..."):
                    upload_success, upload_message = upload_to_django(
                        st.session_state['po_number'],
                        st.session_state.get('party_code_value', ''),
                        st.session_state.get('po_date', ''),
                        st.session_state.get('po_expiry_date', '')
                    )
                    if upload_success:
                        st.success(upload_message)
                    else:
                        st.error(upload_message)
        else:
            st.info("📧 Email & Upload disabled. Create Email_Config.xlsx to enable")
