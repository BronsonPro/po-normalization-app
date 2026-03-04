import pandas as pd
import re


# ------------------ HEADER EXTRACTION ------------------

def extract_po_header(raw):

    party_name = "BIG BASKET"
    po_no = ""
    po_date = ""
    po_expiry = ""
    shipping_address = ""
    gst_no = ""

    for i in range(len(raw)):
        row = raw.iloc[i].astype(str).tolist()

        for cell in row:
            cell_str = str(cell).strip()
            if "PO Number:" in cell_str:
                po_no = cell_str.split("PO Number:")[-1].strip()
            if "PO Date:" in cell_str:
                po_date = cell_str.split("PO Date:")[-1].strip()
            if "PO Expiry date:" in cell_str or "PO Expiry Date:" in cell_str:
                po_expiry = cell_str.split(":", 1)[-1].strip()

        # Find Delivery Address section - "Delivery Address" is in col7, "Warehouse Address" in col0
        col7 = str(row[7]).strip() if len(row) > 7 else ""
        if "Warehouse Address" in str(row[0]) or "Delivery Address" in col7:
            addr_parts = []
            for k in range(i + 1, min(i + 8, len(raw))):
                addr_row = raw.iloc[k].astype(str).tolist()
                val = str(addr_row[7]).strip() if len(addr_row) > 7 else ""
                if val and val != "nan":
                    if "GSTIN" in val:
                        gst_match = re.search(r'[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[A-Z0-9]{3}', val)
                        if gst_match:
                            gst_no = gst_match.group(0)
                        break
                    elif "Supplier" in val:
                        break
                    else:
                        addr_parts.append(val)
            shipping_address = " ".join(addr_parts)

    return {
        "Party Name": party_name,
        "PO No": po_no,
        "PO Date": po_date,
        "PO Expiry Date": po_expiry,
        "Shipping Address": shipping_address,
        "GST #": gst_no,
    }


# ------------------ PRODUCTS EXTRACTION ------------------

def extract_products(raw):

    # Find header row (where S.No and EAN/UPC Code appear)
    header_row = None
    for i in range(len(raw)):
        row = raw.iloc[i].astype(str).str.strip().tolist()
        if "S.No" in row and any("EAN" in c for c in row):
            header_row = i
            break

    if header_row is None:
        raise Exception("Product table header not found in BigBasket PO")

    # Read from header row
    df = raw.iloc[header_row:].copy()
    df.columns = df.iloc[0].astype(str).str.strip()
    df = df.iloc[1:].reset_index(drop=True)

    # Keep only valid product rows (S.No is numeric)
    df = df[pd.to_numeric(df["S.No"], errors="coerce").notna()].copy()
    df = df.reset_index(drop=True)

    # Build normalized products dataframe
    products = pd.DataFrame()
    products["Sr #"] = range(1, len(df) + 1)
    products["EAN"] = df["EAN/UPC Code"].astype(str).str.strip().str.replace(".0", "", regex=False)
    products["Product Name"] = df["Description"].astype(str).str.strip()
    products["HSN Code"] = df["HSN Code"].astype(str).str.strip().str.replace(".0", "", regex=False)
    products["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    products["MRP"] = pd.to_numeric(df["MRP"], errors="coerce").fillna(0).round(2)
    # Base Rate = Landing Cost (used for master comparison)
    products["Base Rate"] = pd.to_numeric(df["Landing Cost"], errors="coerce").fillna(0).round(2)
    products["GST %"] = pd.to_numeric(df["GST%"], errors="coerce").fillna(0).round(2)
    # Total = Basic Cost x Qty x (1 + GST%)
    products["Total"] = pd.to_numeric(df["Total Value"], errors="coerce").fillna(0).round(2)

    return products, header_row


# ------------------ SUMMARY ------------------

def extract_summary(excel_path, products):
    """Extract summary - sum GST Amount from product rows directly"""

    from openpyxl import load_workbook

    grand_total = round(products["Total"].sum(), 2)
    total_tax = 0.0

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        # Find header row and GST Amount column index
        gst_amt_col = None
        header_row_num = None

        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value).strip() == "GST Amount":
                    gst_amt_col = cell.column
                    header_row_num = cell.row
                    break
            if gst_amt_col:
                break

        # Sum GST Amount from all product rows (rows after header until empty S.No)
        if gst_amt_col and header_row_num:
            for r in range(header_row_num + 1, ws.max_row + 1):
                sno = ws.cell(row=r, column=1).value
                # Stop at non-numeric S.No (Total row or empty)
                try:
                    float(str(sno))
                except:
                    break
                val = ws.cell(row=r, column=gst_amt_col).value
                if val is not None:
                    try:
                        total_tax += float(val)
                    except:
                        pass
            total_tax = round(total_tax, 2)

    except Exception as e:
        print(f"Summary extraction error: {e}")

    return {
        "Total Base Value": "",
        "Total Tax": f"{total_tax:.2f}",
        "Grand Total": f"{grand_total:.2f}",
    }
# ================== PUBLIC FUNCTION ==================

def convert_pdf_to_excel(excel_input_path, output_excel_path):
    """
    BigBasket PO is already Excel.
    Converts it to the normalized format used by all other parsers.
    """

    # Read raw Excel
    raw = pd.read_excel(excel_input_path, header=None)

    # Extract all sections
    header_data = extract_po_header(raw)
    products, _ = extract_products(raw)
    summary_data = extract_summary(excel_input_path, products)

    # Write normalized output Excel
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        row_offset = 0

        # Header section (Field | Value rows)
        header_df = pd.DataFrame({
            "Field": list(header_data.keys()),
            "Value": list(header_data.values()),
        })
        header_df.to_excel(writer, index=False, startrow=row_offset, header=False)
        row_offset += len(header_df) + 2

        # Products table
        products.to_excel(writer, index=False, startrow=row_offset)
        row_offset += len(products) + 2

        # Summary section
        summary_df = pd.DataFrame({
            "Field": list(summary_data.keys()),
            "Value": list(summary_data.values()),
        })
        summary_df.to_excel(writer, index=False, startrow=row_offset, header=False)