import pdfplumber
import pandas as pd
import re


# ------------------ HEADER EXTRACTION ------------------

def extract_po_header(pdf_path):
    party_name = "Sliksync"
    po_no = ""
    po_date = ""
    po_expiry = ""
    shipping_address = ""
    gst_no = ""

    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""
        tables = pdf.pages[0].extract_tables()

    # PO No
    m = re.search(r'PO#\s*(SLIKPO#\d+)', text)
    if m:
        po_no = m.group(1).strip()

    # Order Date, PO Expiry, Shipping Address from tables
    for table in tables:
        for row in table:
            r = [str(c).strip() if c else "" for c in row]
            for i, cell in enumerate(r):
                if "ORDER" in cell and "DATE" in cell and i + 1 < len(r):
                    po_date = r[i + 1].replace("\n", " ").strip()
                if "PO Expiry" in cell and i + 1 < len(r):
                    po_expiry = r[i + 1].strip()
                if "SHIPPED TO" in cell and "ADDRESS" in cell and i + 1 < len(r):
                    addr = r[i + 1].replace("\n", " ").strip()
                    pin_match = re.search(r'\d{6}', addr)
                    if pin_match:
                        shipping_address = pin_match.group(0)

    # Slikk GSTIN (Karnataka = state code 29)
    gstin_matches = re.findall(r'\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1})\b', text)
    for g in gstin_matches:
        if g.startswith("29"):
            gst_no = g
            break

    return {
        "Party Name": party_name,
        "PO No": po_no,
        "PO Date": po_date,
        "PO Expiry Date": po_expiry,
        "Shipping Address": shipping_address,
        "GST #": gst_no,
    }


# ------------------ LINE ITEMS EXTRACTION ------------------

def extract_line_items(pdf_path):
    # Pass 1: collect all table rows and pending continuation rows per page
    # Pass 2: collect text-only rows (missed by table parser)
    # Then merge by S.No

    table_items = {}   # sno -> item dict
    text_items = {}    # sno -> partial item dict (data only, no name/sku)
    continuation = {}  # sno -> {name, sku} from continuation rows

    with pdfplumber.open(pdf_path) as pdf:
        prev_sno = 0  # track last seen S.No for continuation rows

        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()

            for table in tables:
                if not table:
                    continue
                header_row = [str(c).replace("\n", " ").strip().lower() if c else "" for c in table[0]]
                if "s.no" not in " ".join(header_row) or "sku" not in " ".join(header_row):
                    continue

                for row in table[1:]:
                    r = [str(c).replace("\n", "").strip() if c else "" for c in row]

                    sno_raw        = r[0]
                    name           = r[1]
                    sku_raw        = r[2]
                    hsn            = r[3]
                    mrp            = r[4]
                    qty_str        = r[6]
                    purchase_price = r[7]
                    item_value     = r[8]
                    total_tax      = r[9]
                    total_amt      = r[10] if len(r) > 10 else ""

                    sku_digits = re.sub(r'\D', '', sku_raw)
                    sku = sku_digits[:13] if len(sku_digits) >= 13 else sku_digits

                    sno_int = int(sno_raw) if sno_raw.isdigit() else 0

                    # Continuation row: has name/sku but no data
                    if not hsn and not qty_str:
                        if sno_int > 0:
                            prev_sno = sno_int
                        # Store as continuation for the next missing sno
                        target_sno = sno_int if sno_int > 0 else prev_sno + 1
                        continuation[target_sno] = {
                            "name": name if name else "",
                            "sku": sku if sku else "",
                        }
                        if sno_int == 0:
                            prev_sno = target_sno
                        continue

                    if sno_int > 0:
                        prev_sno = sno_int

                    try:
                        qty = int(float(qty_str)) if qty_str else 0
                    except:
                        qty = 0

                    if qty <= 0:
                        continue

                    try:
                        gst_pct = round((float(total_tax) / float(item_value)) * 100) if item_value and float(item_value) > 0 else 18
                    except:
                        gst_pct = 18

                    try:
                        mrp_val   = float(mrp) if mrp else 0
                        base_val  = float(purchase_price) if purchase_price else 0
                        total_val = float(total_amt) if total_amt else 0
                    except:
                        mrp_val = base_val = total_val = 0

                    table_items[sno_int] = {
                        "EAN": sku,
                        "Product Name": name,
                        "HSN Code": hsn,
                        "Quantity": qty,
                        "MRP": mrp_val,
                        "Base Rate": base_val,
                        "GST %": gst_pct,
                        "Total": total_val,
                    }

            # Fallback: extract rows from page text that the table missed
            text = page.extract_text() or ""
            text_rows = re.findall(
                r'^(\d+)\s+(\d{8})\s+([\d.]+)\s+([\d.]+)\s+(\d+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)',
                text, re.MULTILINE
            )
            for tr in text_rows:
                sno_t = int(tr[0])
                if sno_t in table_items:
                    continue
                hsn_t      = tr[1]
                mrp_t      = float(tr[2])
                qty_t      = int(tr[4])
                base_t     = float(tr[5])
                item_val_t = float(tr[6])
                tax_t      = float(tr[7])
                total_t    = float(tr[8])
                try:
                    gst_t = round((tax_t / item_val_t) * 100) if item_val_t > 0 else 18
                except:
                    gst_t = 18

                text_items[sno_t] = {
                    "HSN Code": hsn_t,
                    "Quantity": qty_t,
                    "MRP": mrp_t,
                    "Base Rate": base_t,
                    "GST %": gst_t,
                    "Total": total_t,
                }

    # Merge text_items with continuation data
    for sno, data in text_items.items():
        cont = continuation.get(sno, {})
        table_items[sno] = {
            "EAN": cont.get("sku", ""),
            "Product Name": cont.get("name", ""),
            **data,
        }

    # Also apply continuation to table_items missing name/sku
    for sno, cont in continuation.items():
        if sno in table_items:
            if not table_items[sno].get("EAN") and cont.get("sku"):
                table_items[sno]["EAN"] = cont["sku"]
            if not table_items[sno].get("Product Name") and cont.get("name"):
                table_items[sno]["Product Name"] = cont["name"]

    # Build sorted list
    items = []
    for sno in sorted(table_items.keys()):
        item = table_items[sno]
        item["Sr #"] = len(items) + 1
        items.append(item)

    return pd.DataFrame(items)


# ------------------ SUMMARY EXTRACTION ------------------

def extract_summary(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        tables = pdf.pages[0].extract_tables()

    taxable_val = ""
    tax_amt = ""
    total_amt = ""

    for table in tables:
        for i, row in enumerate(table):
            r = [str(c).strip() if c else "" for c in row]
            # Header row: Taxable Value | Tax Amount | Total Amount
            if "Taxable Value" in r and "Tax Amount" in r and "Total Amount" in r:
                # Next row has values
                if i + 1 < len(table):
                    vals = [str(c).strip() if c else "" for c in table[i + 1]]
                    taxable_val = vals[0].replace(" INR", "").replace(",", "").strip()
                    tax_amt     = vals[1].replace(" INR", "").replace(",", "").strip()
                    total_amt   = vals[2].replace(" INR", "").replace(",", "").strip()
                break

    return {
        "Total Base Value": taxable_val,
        "Total Tax": tax_amt,
        "Grand Total": total_amt,
    }


# ================== PUBLIC FUNCTION ==================

def convert_pdf_to_excel(pdf_path, output_excel_path):

    header_data = extract_po_header(pdf_path)
    products = extract_line_items(pdf_path)

    if products.empty:
        raise Exception("No line items found in Slikk PO")

    summary_data = extract_summary(pdf_path)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    row_offset = 1

    for field, value in header_data.items():
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=value)
        row_offset += 1

    row_offset += 2

    headers = ["Sr #", "EAN", "Product Name", "HSN Code", "Quantity", "MRP", "Base Rate", "GST %", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row_offset, column=col, value=header)
    row_offset += 1

    for _, row in products.iterrows():
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col)
            value = row[header]
            if header in ("EAN", "HSN Code"):
                cell.value = str(value)
                cell.number_format = '@'
            else:
                cell.value = value
        row_offset += 1

    row_offset += 2

    for field, value in summary_data.items():
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=value)
        row_offset += 1

    wb.save(output_excel_path)
