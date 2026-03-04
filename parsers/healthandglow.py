import pdfplumber
import pandas as pd
import re


# ------------------ HEADER EXTRACTION ------------------

def extract_po_header(pdf_path):

    party_name = "Health & Glow"
    po_no = ""
    po_date = ""
    po_expiry = ""
    shipping_address = ""
    gst_no = ""

    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""

    for line in text.split("\n"):

        # PO No : 6367126
        if "PO No :" in line or "PO No:" in line:
            m = re.search(r'PO No\s*:\s*(\d+)', line)
            if m:
                po_no = m.group(1).strip()

        # PO Date : 19-12-2025
        if "PO Date :" in line or "PO Date:" in line:
            m = re.search(r'PO Date\s*:\s*([\d\-]+)', line)
            if m:
                po_date = m.group(1).strip()

        # Expiry Date : 01-01-2026
        if "Expiry Date :" in line or "Expiry Date:" in line:
            m = re.search(r'Expiry Date\s*:\s*([\d\-]+)', line)
            if m:
                po_expiry = m.group(1).strip()

    # Shipping Address and GST - extract from Buyer Details section
    # Buyer section comes AFTER supplier section
    buyer_section = ""
    supplier_section = ""
    lines = text.split("\n")
    
    in_buyer = False
    for line in lines:
        if "Supplier Details" in line:
            in_buyer = False
        if "Buyer Details" in line:
            in_buyer = True
        if in_buyer:
            buyer_section += line + " "
        else:
            supplier_section += line + " "
    
    # Extract pincode from buyer section (last 6-digit number)
    buyer_pins = re.findall(r'Pincode\s*:(\d{6})', buyer_section)
    if buyer_pins:
        shipping_address = buyer_pins[-1]  # Last pincode is buyer's
    
    # GST from buyer section (last GSTIN in buyer section is buyer's)
    buyer_gst = re.findall(r'GSTIN\s*:\s*([A-Z0-9]{15})', buyer_section)
    if buyer_gst:
        gst_no = buyer_gst[-1]  # Last GSTIN in buyer section is the buyer's

    return {
        "Party Name": party_name,
        "PO No": po_no,
        "PO Date": po_date,
        "PO Expiry Date": po_expiry,
        "Shipping Address": shipping_address,
        "GST #": gst_no,
    }


# ------------------ LINE ITEMS EXTRACTION ------------------

def extract_line_items(pdf_path):

    all_lines = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_lines.extend(text.split("\n"))

    items = []

    i = 0
    while i < len(all_lines):
        line = all_lines[i].strip()

        # Match lines starting with serial number followed by product code
        # Example: "1 578087 8652365478898 85437093 1 211.86 GST 18% 38.14 250.00..."
        m = re.match(
            r'^(\d+)\s+'                    # Serial number
            r'(\d{6})\s+'                   # Product Code
            r'(\d{13})\s+'                  # EAN Code
            r'(\d{8})\s+'                   # HSN Code
            r'(\d+)\s+'                     # Qty
            r'([\d.]+)\s+'                  # Base Cost
            r'GST\s+([\d]+)%',              # GST percentage
            line
        )

        if m:
            sr_no = m.group(1)
            product_code = m.group(2)
            ean = m.group(3)
            hsn = m.group(4)
            qty = int(m.group(5))
            base_cost = float(m.group(6))
            gst_pct = float(m.group(7))

            # Extract MRP and Total from the line
            # Pattern: "... 75.00% 0.00% .00 325 619.71"
            # MRP is second-to-last number, Total is last number
            parts = line.split()
            # Find numeric values at the end
            numbers = []
            for p in parts[-10:]:
                try:
                    numbers.append(float(p))
                except:
                    pass
            
            if len(numbers) >= 2:
                total = numbers[-1]
                mrp = numbers[-2]
            else:
                mrp = 0.0
                total = 0.0

            # Product name is on the next line
            product_name = ""
            if i + 1 < len(all_lines):
                next_line = all_lines[i + 1].strip()
                # Product name line doesn't start with a number
                if next_line and not re.match(r'^\d+\s+\d{6}', next_line) and "IGST" not in next_line:
                    product_name = next_line

            items.append({
                "Sr #": int(sr_no),
                "EAN": ean,
                "Product Name": product_name,
                "HSN Code": hsn,
                "Quantity": qty,
                "MRP": mrp,
                "Base Rate": base_cost,
                "GST %": gst_pct,
                "Total": total,
            })

        i += 1

    return pd.DataFrame(items)


# ------------------ SUMMARY EXTRACTION ------------------

def extract_summary(pdf_path):

    grand_total = 0.0
    total_tax = 0.0

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                # PO Total Value : 12024.49
                m = re.search(r'PO Total Value\s*:\s*([\d,]+\.?\d*)', line)
                if m:
                    grand_total = float(m.group(1).replace(",", ""))
                
                # IGST : 1834.24
                m = re.search(r'IGST\s*:\s*([\d,]+\.?\d*)', line)
                if m:
                    total_tax = float(m.group(1).replace(",", ""))

    # Calculate base value
    total_base = round(grand_total - total_tax, 2)

    return {
        "Total Base Value": f"{total_base:.2f}",
        "Total Tax": f"{total_tax:.2f}",
        "Grand Total": f"{grand_total:.2f}",
    }


# ================== PUBLIC FUNCTION ==================

def convert_pdf_to_excel(pdf_path, output_excel_path):

    header_data = extract_po_header(pdf_path)
    products = extract_line_items(pdf_path)

    if products.empty:
        raise Exception("No line items found in Health & Glow PO")

    summary_data = extract_summary(pdf_path)

    # Write to Excel with proper formatting
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    row_offset = 1
    
    # Header section
    for field, value in header_data.items():
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=value)
        row_offset += 1
    
    row_offset += 2
    
    # Products table
    headers = ["Sr #", "EAN", "Product Name", "HSN Code", "Quantity", "MRP", "Base Rate", "GST %", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row_offset, column=col, value=header)
    row_offset += 1
    
    # Products data
    for _, row in products.iterrows():
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col)
            value = row[header]
            
            # Format EAN as text
            if header == "EAN":
                cell.value = str(value)
                cell.number_format = '@'
            else:
                cell.value = value
        row_offset += 1
    
    row_offset += 2
    
    # Summary section
    for field, value in summary_data.items():
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=value)
        row_offset += 1
    
    wb.save(output_excel_path)