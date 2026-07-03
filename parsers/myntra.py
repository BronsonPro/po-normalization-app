import pdfplumber
import pandas as pd
import re

# ------------------ HEADER EXTRACTION ------------------

def extract_po_header(pdf_path):
    party_name = "Myntra"
    po_no = ""
    po_date = ""
    po_expiry = ""
    shipping_address = ""
    gst_no = ""

    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""

        for line in text.split("\n"):
            if "PO #:" in line:
                m = re.search(r'PO #:\s*([A-Z0-9\-]+)', line)
                if m:
                    po_no = m.group(1).strip()

            if "PO Approved Date:" in line:
                m = re.search(r'PO Approved Date:\s*([\d\-]+)', line)
                if m:
                    date_str = m.group(1).strip()
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(date_str, "%Y-%m-%d")
                        po_date = dt.strftime("%d-%m-%Y")
                    except:
                        po_date = date_str

            if "Estimated Shipment Date:" in line:
                m = re.search(r'Estimated Shipment Date:\s*([\d/]+)', line)
                if m:
                    date_str = m.group(1).strip()
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(date_str, "%d/%m/%Y")
                        po_expiry = dt.strftime("%d-%m-%Y")
                    except:
                        po_expiry = date_str

            if "GSTIN#" in line:
                m = re.search(r'GSTIN#\s*([A-Z0-9]{15})', line)
                if m:
                    gst_no = m.group(1).strip()

        # Extract pincode
        pin_match = re.search(r'\b(\d{6})\b', text)
        if pin_match:
            shipping_address = pin_match.group(1)

    return {
        "Party Name": party_name,
        "PO No": po_no,
        "PO Date": po_date,
        "PO Expiry Date": po_expiry,
        "Shipping Address": shipping_address,
        "GST #": gst_no,
    }

# ------------------ LINE ITEMS EXTRACTION ------------------

def extract_line_items(pdf_path):
    items = []
    
    with pdfplumber.open(pdf_path) as pdf:
        # Extract table from page 1 (diagnostic showed table headers exist)
        tables = pdf.pages[0].extract_tables()
        
        if not tables:
            return pd.DataFrame(items)
        
        # Find the product table (the one with column headers)
        header_row = None
        data_table = None
        
        for table in tables:
            if not table:
                continue
            # Look for header row containing SKU Code
            for row in table:
                if row and any('SKU' in str(cell) for cell in row):
                    header_row = row
                    data_table = table
                    break
            if header_row:
                break
        
        if not header_row:
            # Fallback to text-based parsing
            return extract_line_items_from_text(pdf_path)
        
        # Clean header names
        headers = [str(cell).strip().replace('\n', ' ') for cell in header_row]
        
        # Find column indices by name
        def find_col(keywords):
            for i, h in enumerate(headers):
                if any(kw.lower() in h.lower() for kw in keywords):
                    return i
            return -1
        
        sku_col = find_col(['SKU'])
        hsn_col = find_col(['HSN'])
        product_col = find_col(['Article', 'Name', 'Product'])
        ean_col = find_col(['Article Number', 'EAN', 'Vendor Article Number'])
        qty_col = find_col(['Quantity', 'Qty'])
        mrp_col = find_col(['MRP'])
        
        # For base rate and GST, we need to check what columns exist
        cgst_pct_col = find_col(['CGST %', 'CGST%'])
        sgst_pct_col = find_col(['SGST %', 'SGST%'])
        igst_pct_col = find_col(['IGST %', 'IGST%'])
        
        # Find rate columns - could be "Landed Cost", "Base Rate", etc.
        rate_cols = []
        for i, h in enumerate(headers):
            if any(kw in h.lower() for kw in ['cost', 'rate', 'price']) and 'mrp' not in h.lower():
                rate_cols.append(i)
        
        total_col = find_col(['Total', 'Amount'])
        
        # Now extract from text since table structure doesn't preserve all data properly
        # Use table only to understand structure, extract from text
        return extract_line_items_from_text(pdf_path)

def _reconstruct_ean_from_words(page, sku_code, x_min=140, x_max=207, top_window=25):
    """Fallback EAN reconstruction for rows whose Product Name/EAN cell wraps
    across a page break. In that situation pdfplumber's extract_text() can
    merge the wrapped sub-lines into one garbled line, scrambling the EAN's
    digit order. This works around it by reading pdfplumber's word-level
    bounding boxes (which stay individually accurate even when line assembly
    doesn't) for digit-only words in the Product Name/EAN column band, sorted
    by true (top, x0) position - the genuine reading order."""
    try:
        words = page.extract_words()
    except Exception:
        return None

    sku_words = [w for w in words if w['text'] == sku_code]
    if not sku_words:
        return None
    top_center = sku_words[0]['top']

    candidates = [
        w for w in words
        if x_min <= w['x0'] <= x_max
        and abs(w['top'] - top_center) <= top_window
        and w['text'].isdigit()
    ]
    if not candidates:
        return None

    groups = {}
    for w in candidates:
        groups.setdefault(round(w['top'], 1), []).append(w)

    result = []
    for t in sorted(groups.keys()):
        group = sorted(groups[t], key=lambda w: w['x0'])
        kept = []
        for w in group:
            if kept and w['x0'] < kept[-1]['x1'] - 0.5:
                # Overlapping bbox - likely a duplicated glyph; keep the longer read
                if len(w['text']) > len(kept[-1]['text']):
                    kept[-1] = w
            else:
                kept.append(w)
        result.extend(w['text'] for w in kept)

    digits = "".join(result)
    return digits if 13 <= len(digits) <= 14 else None


def extract_line_items_from_text(pdf_path):
    """Extract using text with intelligent column detection"""
    items = []
    debug_info = []  # Collect debug info to show in UI
    
    with pdfplumber.open(pdf_path) as pdf:
        lines = []
        line_page_idx = []  # parallel array: which page (index into pdf.pages) each line came from
        for p_idx, page in enumerate(pdf.pages):
            page_text = page.extract_text() or ""
            page_lines = page_text.split('\n')
            lines.extend(page_lines)
            line_page_idx.extend([p_idx] * len(page_lines))
        
        # Find header line - check both text and look for tax keywords
        header_line = None
        for line in lines:
            # Look for header containing tax-related keywords
            if any(kw in line for kw in ['SKU', 'HSN', 'CGST', 'SGST', 'IGST']):
                header_line = line
                break
        
        # Detect if IGST or CGST+SGST based on header keywords
        is_igst = None
        is_cgst_sgst = None
        
        if header_line:
            # Check for CGST and SGST together (intrastate)
            if 'CGST' in header_line and 'SGST' in header_line:
                is_cgst_sgst = True
                is_igst = False
            # Check for IGST (interstate)
            elif 'IGST' in header_line:
                is_igst = True
                is_cgst_sgst = False
        
        debug_info.append(f"Header line: {header_line[:150] if header_line else 'NOT FOUND'}")
        debug_info.append(f"From header - IGST: {is_igst}, CGST+SGST: {is_cgst_sgst}")
        
        # If header detection failed, detect from first product line
        if is_igst is None and is_cgst_sgst is None:
            for line in lines:
                if 'BNPL' in line:
                    parts = line.strip().split()
                    # CGST+SGST has duplicated percentages (9.00 appears twice)
                    # Check if parts[-3] == parts[-5] (SGST% == CGST%)
                    try:
                        if len(parts) >= 10:
                            val1 = float(parts[-3])
                            val2 = float(parts[-5])
                            if val1 == val2 and val1 < 50:  # Same percentage (likely CGST == SGST)
                                is_cgst_sgst = True
                                is_igst = False
                                debug_info.append(f"Auto-detected CGST+SGST (found matching values: {val1} == {val2})")
                            else:
                                is_igst = True
                                is_cgst_sgst = False
                                debug_info.append(f"Auto-detected IGST (values don't match: {val1} != {val2})")
                            break
                    except:
                        pass
        
        # Final fallback
        if is_igst is None:
            is_igst = False
            is_cgst_sgst = True
            debug_info.append("Defaulting to CGST+SGST format")
        
        reconstructed_eans = []
        
        for idx, line in enumerate(lines):
            if 'BNPL' not in line:
                continue
            
            parts = line.strip().split()
            
            if len(parts) < 10:
                continue
            
            try:
                # Find SKU starting with BNPL
                sku_idx = -1
                for i, part in enumerate(parts):
                    if part.startswith('BNPL'):
                        sku_idx = i
                        break
                
                if sku_idx == -1:
                    continue
                
                sku_code = parts[sku_idx]
                
                # HSN is right after SKU
                hsn = parts[sku_idx + 1] if sku_idx + 1 < len(parts) else ""
                
                # Determine the numeric tail (Qty/MRP/Rate/Tax/Total) FIRST using fixed
                # offsets from the end of the line. These fields are reliable even when
                # the Product Name / EAN columns get scrambled by page-break text wrapping,
                # so we use them to compute where the Style ID sits and bound the EAN
                # search to stop before it - this stops the parser from ever mistaking
                # the Style ID for the EAN.
                if is_igst or (not is_cgst_sgst and len(parts) < sku_idx + 20):
                    style_id_idx = len(parts) - 8
                else:
                    style_id_idx = len(parts) - 10
                ean_search_end = max(sku_idx + 2, min(style_id_idx, len(parts)))
                
                # Find EAN - look for 8-digit or longer number, but only before the Style ID
                # EAN might be split across 2 lines
                ean = ""
                ean_idx = -1
                for i in range(sku_idx + 2, ean_search_end):
                    if parts[i].isdigit() and len(parts[i]) >= 8:
                        ean = parts[i]
                        ean_idx = i
                        break
                
                # Fallback 1: the EAN's digits may have been scrambled into isolated
                # fragments (happens when this row's Product Name/EAN cell wraps
                # across a page break and pdfplumber's line assembly interleaves it
                # with the row's data columns). Simply concatenating the stray digit
                # tokens in "parts" order is NOT reliable here - testing showed it
                # can reproduce the digits in the wrong order. Instead, reconstruct
                # using each word's actual (top, x0) position on the page.
                ean_reconstructed = False
                if not ean:
                    page_for_line = pdf.pages[line_page_idx[idx]]
                    reconstructed = _reconstruct_ean_from_words(page_for_line, sku_code)
                    if reconstructed:
                        ean = reconstructed
                        ean_idx = ean_search_end  # product name slice below just uses this as an end bound
                        ean_reconstructed = True
                        debug_info.append(f"Word-position reconstructed EAN: {ean}")
                
                # Fallback 2: the wrap may instead start on the PRECEDING line/page
                # (e.g. the last item on a page whose Product Name/EAN text spills
                # onto the following page, leaving only a short digit fragment on
                # the SKU's own line). Look a few lines backward - stopping if we
                # hit the previous item's SKU line - for an 8+ digit prefix, then
                # complete it using a short fragment from the SKU's own line.
                if not ean:
                    prefix = ""
                    for back in range(1, 6):
                        back_idx = idx - back
                        if back_idx < 0:
                            break
                        back_line = lines[back_idx].strip()
                        if 'BNPL' in back_line:
                            break  # ran into the previous item - stop
                        back_parts = back_line.split()
                        found_prefix = next((p for p in back_parts if p.isdigit() and len(p) >= 8), None)
                        if found_prefix:
                            prefix = found_prefix
                            break
                    if prefix:
                        for i in range(sku_idx + 2, ean_search_end):
                            p = parts[i]
                            if p.isdigit() and 3 <= len(p) <= 6:
                                candidate = prefix + p
                                if 13 <= len(candidate) <= 14:
                                    ean = candidate
                                    ean_idx = i + 1
                                    ean_reconstructed = True
                                    debug_info.append(f"Backward-reconstructed EAN: {prefix} + {p} = {ean}")
                                    break
                
                # If EAN found and it's less than 13 digits, check next line for continuation
                if ean and len(ean) < 13 and idx + 1 < len(lines):
                    next_line = lines[idx + 1].strip()
                    if next_line and not next_line.startswith('BNPL'):
                        next_parts = next_line.split()
                        # Look for first numeric part that could be EAN continuation
                        for np in next_parts:
                            if np.isdigit() and 2 <= len(np) <= 6:
                                potential_ean = ean + np
                                # Valid EAN is 13 or 14 digits
                                if 13 <= len(potential_ean) <= 14:
                                    ean = potential_ean
                                    debug_info.append(f"Multi-line EAN detected: {parts[ean_idx]} + {np} = {ean}")
                                    break
                
                if not ean:
                    continue
                
                # Product name is between HSN and EAN on current line
                product_name_parts = parts[sku_idx + 2:ean_idx]
                
                # Check next 2-3 lines for continuation of product name (common in Myntra POs)
                # Product name often wraps to multiple lines
                for line_offset in range(1, 4):  # Check next 3 lines
                    if idx + line_offset >= len(lines):
                        break
                    next_line = lines[idx + line_offset].strip()
                    if not next_line or next_line.startswith('BNPL'):
                        break
                    
                    next_parts = next_line.split()
                    added_any = False
                    
                    # Add words from next line until we hit numbers/codes/colors
                    for p in next_parts:
                        # Stop conditions:
                        # - Pure digits with 6+ chars (EAN, Style ID)
                        if p.isdigit() and len(p) >= 6:
                            break
                        # - Common color/size keywords
                        if p.upper() in ['ONESIZE', 'PACK', 'BROWN', 'GOLDEN', 'MULTI', 'SILVER', 'BLACK', 'WHITE', 'RED', 'BLUE', 'GREEN', 'PINK', 'YELLOW', 'ORANGE', 'PURPLE', 'GREY', 'GRAY']:
                            break
                        # - Single letter or digit
                        if len(p) == 1:
                            break
                        
                        # Add if it contains letters (it's a word)
                        if any(c.isalpha() for c in p):
                            product_name_parts.append(p)
                            added_any = True
                        # If it's a short number (2-5 digits), might be part of EAN, stop
                        elif p.isdigit() and 2 <= len(p) <= 5:
                            break
                        else:
                            break
                    
                    # If we didn't add anything from this line, stop checking further lines
                    if not added_any:
                        break
                
                product_name = " ".join(product_name_parts).strip()
                debug_info.append(f"Product name captured: {product_name}")
                
                # Now intelligently find numeric columns from the end
                # Total is always last
                total = float(parts[-1])
                
                debug_info.append(f"\n=== Product: {sku_code} ===")
                debug_info.append(f"Total parts: {len(parts)}")
                debug_info.append(f"Last 10 parts: {parts[-10:]}")
                
                # Determine format and extract accordingly
                if is_igst or (not is_cgst_sgst and len(parts) < sku_idx + 20):
                    # IGST format: ...Qty MRP Rate1 Rate2 IGST% IGST_Amt Total
                    debug_info.append("Using IGST format")
                    igst_amt = float(parts[-2])
                    gst_pct = float(parts[-3])
                    base_rate2 = float(parts[-4])
                    base_rate1 = float(parts[-5])
                    mrp = float(parts[-6])
                    qty = int(float(parts[-7]))
                    debug_info.append(f"Qty={qty}, MRP={mrp}, Rate={base_rate1}, GST%={gst_pct}, Total={total}")
                else:
                    # CGST+SGST format: ...Qty MRP Rate1 Rate2 CGST% CGST_Amt SGST% SGST_Amt Total
                    debug_info.append("Using CGST+SGST format")
                    sgst_amt = float(parts[-2])
                    sgst_pct = float(parts[-3])
                    cgst_amt = float(parts[-4])
                    cgst_pct = float(parts[-5])
                    base_rate2 = float(parts[-6])
                    base_rate1 = float(parts[-7])
                    mrp = float(parts[-8])
                    qty = int(float(parts[-9]))
                    gst_pct = cgst_pct + sgst_pct
                    debug_info.append(f"Qty={qty}, MRP={mrp}, Rate={base_rate1}, GST%={gst_pct}, Total={total}")
                
                if qty <= 0 or mrp <= 0 or total <= 0:
                    continue
                
                items.append({
                    "Sr #": len(items) + 1,
                    "EAN": ean,
                    "Product Name": product_name,
                    "HSN Code": hsn,
                    "Quantity": qty,
                    "MRP": mrp,
                    "Base Rate": base_rate1,
                    "GST %": gst_pct,
                    "Total": total,
                })
                if ean_reconstructed:
                    reconstructed_eans.append((sku_code, ean))
                
            except (ValueError, IndexError) as e:
                debug_info.append(f"ERROR parsing: {str(e)}")
                continue
    
    # Flag any EANs that had to be reconstructed (row wrapped across a page
    # break) so they get a manual once-over rather than being silently trusted -
    # these are the rows most likely to need a spot-check.
    if reconstructed_eans:
        import streamlit as st
        st.warning(
            "⚠️ EAN reconstructed for " + str(len(reconstructed_eans)) +
            " item(s) whose row wrapped across a page break - please verify: " +
            ", ".join(f"{sku} → {ean}" for sku, ean in reconstructed_eans)
        )
    
    # Show debug in Streamlit
    import streamlit as st
    with st.expander("🔍 Myntra Parser Debug Info"):
        st.text("\n".join(debug_info))
    
    return pd.DataFrame(items)

# ------------------ SUMMARY EXTRACTION ------------------

def extract_summary(pdf_path):
    grand_total = 0.0
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            m = re.search(r'Grand Total:\s*([\d,]+\.?\d*)', text)
            if m:
                grand_total = float(m.group(1).replace(",", ""))
                break
    
    return grand_total

# ================== PUBLIC FUNCTION ==================

def convert_pdf_to_excel(pdf_path, output_excel_path):
    header_data = extract_po_header(pdf_path)
    products = extract_line_items(pdf_path)

    if products.empty:
        raise Exception("No line items found in Myntra PO")

    grand_total = extract_summary(pdf_path)
    total_base = round((products["Base Rate"] * products["Quantity"]).sum(), 2)
    total_tax = round(grand_total - total_base, 2)

    summary_data = {
        "Total Base Value": f"{total_base:.2f}",
        "Total Tax": f"{total_tax:.2f}",
        "Grand Total": f"{grand_total:.2f}",
    }

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    row_offset = 1

    for field, value in header_data.items():
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=value)
        row_offset += 1

    row_offset += 2

    headers = ["Sr #", "EAN", "Product Name", "HSN Code", "Quantity", "MRP", "Base Rate", "GST %", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row_offset, column=col, value=header)

    row_offset += 1

    for _, row in products.iterrows():
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col)
            value = row[header]

            if header == "EAN":
                cell.value = str(value)
                cell.number_format = '@'
            else:
                cell.value = value

        row_offset += 1

    row_offset += 2

    for field, value in summary_data.items():
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=value)
        row_offset += 1

    wb.save(output_excel_path)
